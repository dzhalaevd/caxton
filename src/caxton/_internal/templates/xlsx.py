from __future__ import annotations

import dataclasses
from collections.abc import Iterator, Sequence
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName

from caxton._internal.block_paths import iter_blocks_with_paths
from caxton._internal.compiler import SpreadsheetCompiler
from caxton._internal.const import SPREADSHEET_MAX_COLUMNS, SPREADSHEET_MAX_ROWS
from caxton._internal.layout import DocumentPlan, plan_document
from caxton._internal.shape import table_needs_preparation
from caxton.core.errors import (
    AmbiguousTemplateRefError,
    IncompatibleTemplateRefError,
    InvalidTemplateRefError,
    MissingTemplateRefError,
    Notification,
    TemplateError,
    UnsupportedFeatureError,
)
from caxton.core.ir import (
    CellAddress,
    CellRange,
    SpreadsheetBlockKind,
    SpreadsheetIR,
    SpreadsheetTableIR,
    SpreadsheetWorksheetIR,
)
from caxton.core.models import (
    Column,
    SpreadsheetBlock,
    SpreadsheetDocument,
    SpreadsheetTable,
    TemplateCompilationResult,
    TemplateContext,
    TemplateRef,
    TemplateRepeat,
    TemplateSpecification,
    iter_tables,
)
from caxton.core.models.extensions import PivotBinding
from caxton.core.protocols import DataSourceInfo

if TYPE_CHECKING:
    from caxton._internal.backends.openpyxl.package import PivotDescriptor


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxDefinedName:
    """Read-only defined-name facts without an OpenPyXL object."""

    name: str
    scope: str | None
    destinations: Sequence[tuple[str, str]]
    valid_range: bool = True

    def __post_init__(self) -> None:
        object.__setattr__(self, "destinations", tuple(self.destinations))


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxNamedRange:
    """Backend-local physical descriptor resolved from a generic reference."""

    reference: str
    sheet: str
    min_row: int
    min_column: int
    max_row: int
    max_column: int
    scope: str | None = None
    namespace: str = dataclasses.field(default="xlsx.named_range", init=False)

    @property
    def rows(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def columns(self) -> int:
        return self.max_column - self.min_column + 1


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxTemplateContext(TemplateContext):
    """XLSX inspection result containing bytes and plain descriptors only."""

    payload: bytes = dataclasses.field(repr=False, default=b"")
    worksheets: Sequence[str] = ()
    defined_names: Sequence[XlsxDefinedName] = ()
    pivots: Sequence[PivotDescriptor] = ()

    def __post_init__(self) -> None:
        TemplateContext.__post_init__(self)
        object.__setattr__(self, "payload", bytes(self.payload))
        object.__setattr__(self, "worksheets", tuple(self.worksheets))
        object.__setattr__(self, "defined_names", tuple(self.defined_names))
        object.__setattr__(self, "pivots", tuple(self.pivots))

    @property
    def pivot_names(self) -> tuple[str, ...]:
        return tuple(pivot.name for pivot in self.pivots)


@dataclasses.dataclass(frozen=True, slots=True)
class XlsxTableTarget:
    """Resolved target for one compiled table in one worksheet IR."""

    reference: str
    worksheet_index: int
    path: str
    range: XlsxNamedRange
    repeat: bool = False
    namespace: str = dataclasses.field(default="xlsx.table_target", init=False)


class XlsxTemplateInspector:
    """Read an XLSX template without mutating or serializing it."""

    def inspect(self, template: TemplateSpecification) -> XlsxTemplateContext:
        if template.format != "xlsx":
            message = f"XLSX inspector cannot inspect {template.format!r}"
            raise TemplateError(message, context={"format": template.format})
        payload, source = _read_source(template.source)
        try:
            workbook = load_workbook(BytesIO(payload), data_only=False)
        except Exception as error:
            message = "Could not inspect the XLSX template"
            raise TemplateError(message, context={"source": source}) from error
        try:
            names = tuple(_defined_names(workbook))
            worksheets = tuple(worksheet.title for worksheet in workbook.worksheets)
        finally:
            workbook.close()
        from caxton._internal.backends.openpyxl.package import (  # noqa: PLC0415
            inspect_pivots,
        )

        return XlsxTemplateContext(
            format="xlsx",
            source=source,
            references=tuple(name.name for name in names),
            payload=payload,
            worksheets=worksheets,
            defined_names=names,
            pivots=inspect_pivots(payload),
        )


class XlsxTemplateCompiler:
    """Resolve generic template targets and lower spreadsheet intent."""

    def compile(  # noqa: C901
        self,
        document: SpreadsheetDocument,
        context: XlsxTemplateContext,
    ) -> TemplateCompilationResult[SpreadsheetIR]:
        anchors: dict[SpreadsheetBlock, CellAddress] = {}
        measurements: dict[SpreadsheetBlock, tuple[int, int]] = {}
        targets: list[XlsxTableTarget] = []
        for worksheet_index, worksheet in enumerate(document.worksheets):
            if worksheet.name not in context.worksheets:
                message = f"Worksheet {worksheet.name!r} is absent from the template"
                raise IncompatibleTemplateRefError(
                    message,
                    context={"worksheet": worksheet.name},
                )
            for block, block_path in iter_blocks_with_paths(worksheet.blocks):
                if not isinstance(block, SpreadsheetTable):
                    continue
                table = block
                if table.into is None:
                    continue
                reference, repeated = _table_reference(table)
                resolved = _resolve_named_range(context, reference, worksheet.name)
                _validate_named_range_bounds(resolved)
                if len(table.columns) > resolved.columns:
                    message = f"Template target {reference!r} is too narrow"
                    raise IncompatibleTemplateRefError(
                        message,
                        context={
                            "available_columns": resolved.columns,
                            "required_columns": len(table.columns),
                            "reference": reference,
                        },
                    )
                _validate_target_semantics(table, reference)
                if not table_needs_preparation(table):
                    _validate_target_height(table, resolved, repeated)
                anchors[table] = CellAddress(
                    row=resolved.min_row,
                    column=resolved.min_column,
                )
                measurements[table] = (max(resolved.rows - 1, 0), resolved.columns)
                targets.append(
                    XlsxTableTarget(
                        reference=reference,
                        worksheet_index=worksheet_index,
                        path=block_path,
                        range=resolved,
                        repeat=repeated,
                    ),
                )
        _validate_target_overlaps(targets)
        _validate_template_plan(
            plan_document(
                document,
                measurements=measurements,
                anchor_overrides=anchors,
            ),
        )
        template = document.template
        if template is None:
            message = "Template compilation requires a template"
            raise RuntimeError(message)
        _validate_extensions(document, context, template.extensions)
        compiled = SpreadsheetCompiler().compile_validated(
            document,
            anchor_overrides=anchors,
            check_overlaps=False,
        )
        compiled = _materialize_target_rows(compiled, targets)
        _validate_compiled_targets(compiled, targets)
        return TemplateCompilationResult(
            document=compiled,
            context=context,
            targets=tuple(targets),
            extensions=template.extensions,
        )


def _read_source(source: str | bytes) -> tuple[bytes, str]:
    if isinstance(source, bytes):
        return source, "bytes"
    try:
        return Path(source).read_bytes(), source
    except OSError as error:
        message = "Could not read the template source"
        raise TemplateError(message, context={"source": source}) from error


def _defined_names(workbook: Workbook) -> Iterator[XlsxDefinedName]:
    for name in workbook.defined_names.values():
        yield _defined_name(name, None)
    for worksheet in workbook.worksheets:
        for name in worksheet.defined_names.values():
            yield _defined_name(name, worksheet.title)


def _defined_name(name: DefinedName, scope: str | None) -> XlsxDefinedName:
    valid = getattr(name, "type", None) == "RANGE"
    try:
        destinations = tuple(name.destinations) if valid else ()
    except (AttributeError, TypeError, ValueError):
        destinations = ()
        valid = False
    return XlsxDefinedName(
        name=str(name.name),
        scope=scope,
        destinations=destinations,
        valid_range=valid,
    )


def _table_reference(table: SpreadsheetTable) -> tuple[str, bool]:
    target = table.into
    if isinstance(target, TemplateRepeat):
        return target.reference.name, True
    if isinstance(target, TemplateRef):
        return target.name, False
    message = "Table target was not normalized"
    raise RuntimeError(message)


def _resolve_named_range(  # noqa: C901, WPS238
    context: XlsxTemplateContext,
    reference: str,
    worksheet: str,
) -> XlsxNamedRange:
    matching = tuple(
        item
        for item in context.defined_names
        if item.name.casefold() == reference.casefold()
    )
    local = tuple(item for item in matching if item.scope == worksheet)
    applicable = local or tuple(item for item in matching if item.scope is None)
    if not applicable:
        message = f"Template reference {reference!r} was not found"
        raise MissingTemplateRefError(
            message,
            context={"reference": reference, "worksheet": worksheet},
        )
    if len(applicable) != 1 or len(applicable[0].destinations) != 1:
        message = f"Template reference {reference!r} is ambiguous"
        raise AmbiguousTemplateRefError(
            message,
            context={"reference": reference, "worksheet": worksheet},
        )
    selected = applicable[0]
    if not selected.valid_range:
        message = f"Template reference {reference!r} is not a cell range"
        raise InvalidTemplateRefError(message, context={"reference": reference})
    target_sheet, coordinates = selected.destinations[0]
    if target_sheet != worksheet:
        message = f"Template reference {reference!r} targets another worksheet"
        raise IncompatibleTemplateRefError(
            message,
            context={
                "declared_worksheet": worksheet,
                "reference": reference,
                "target_worksheet": target_sheet,
            },
        )
    try:
        min_column, min_row, max_column, max_row = range_boundaries(coordinates)
    except (TypeError, ValueError) as error:
        message = f"Template reference {reference!r} has an invalid range"
        raise InvalidTemplateRefError(message) from error
    if min_column is None or min_row is None or max_column is None or max_row is None:
        message = f"Template reference {reference!r} is not a rectangular range"
        raise InvalidTemplateRefError(message)
    return XlsxNamedRange(
        reference=reference,
        sheet=target_sheet,
        min_row=min_row,
        min_column=min_column,
        max_row=max_row,
        max_column=max_column,
        scope=selected.scope,
    )


def _validate_extensions(  # noqa: C901, WPS238
    document: SpreadsheetDocument,
    context: XlsxTemplateContext,
    extensions: Sequence[object],
) -> None:
    tables = tuple(
        table
        for worksheet in document.worksheets
        for table in iter_tables(worksheet.blocks)
    )
    for extension in extensions:
        if not isinstance(extension, PivotBinding):
            continue
        matching = tuple(
            pivot
            for pivot in context.pivots
            if pivot.name.casefold() == extension.target.casefold()
        )
        if not matching:
            message = f"Pivot target {extension.target!r} was not found"
            raise MissingTemplateRefError(
                message,
                context={"pivot": extension.target},
            )
        if len(matching) > 1:
            message = f"Pivot target {extension.target!r} is ambiguous"
            raise AmbiguousTemplateRefError(message)
        if matching[0].cache_definition_part is None:
            message = f"Pivot target {extension.target!r} has no cache definition"
            raise InvalidTemplateRefError(message)
        if not _has_pivot_source(tables, extension):
            source = _pivot_source_name(extension)
            message = f"Pivot source {source!r} was not found"
            raise MissingTemplateRefError(message, context={"source": source})


def _has_pivot_source(
    tables: Sequence[SpreadsheetTable],
    binding: PivotBinding,
) -> bool:
    source = _pivot_source_name(binding).casefold()
    return any(
        (table.name is not None and table.name.casefold() == source)
        or (table.into is not None and _table_reference(table)[0].casefold() == source)
        for table in tables
    )


def _pivot_source_name(binding: PivotBinding) -> str:
    return binding.source.name


def _validate_target_height(
    table: SpreadsheetTable,
    target: XlsxNamedRange,
    repeated: bool,
) -> None:
    source = table.data.source
    row_count = source.row_count if isinstance(source, DataSourceInfo) else None
    if repeated or row_count is None or row_count <= target.rows:
        return
    message = f"Template target {target.reference!r} has too few rows"
    raise IncompatibleTemplateRefError(
        message,
        context={
            "available_rows": target.rows,
            "reference": target.reference,
            "required_rows": row_count,
        },
    )


def _validate_target_semantics(table: SpreadsheetTable, reference: str) -> None:
    unsupported: list[str] = []
    for name, enabled in (
        ("autofilter", table.autofilter),
        ("auto_width", table.auto_width is not None),
        ("footer", table.footer is not None),
        ("freeze_header", table.freeze_header),
        ("header_style", table.header_style is not None),
        ("native_table", table.name is not None),
        ("rules", bool(table.rules)),
        ("style", table.style is not None),
    ):
        if enabled:
            unsupported.append(name)
    if any(_has_target_column_presentation(column) for column in table.columns):
        unsupported.append("column_presentation")
    if not unsupported:
        return
    message = "XLSX target tables do not materialize presentation intent"
    raise UnsupportedFeatureError(
        message,
        context={"features": tuple(unsupported), "reference": reference},
    )


def _has_target_column_presentation(column: Column) -> bool:
    grouping_merge = column.grouping is not None and column.grouping.merge
    return any(
        (
            column.alignment is not None,
            column.auto_width is not None,
            column.display_format is not None,
            column.style_ref is not None,
            column.width_hint is not None,
            grouping_merge,
        ),
    )


def _validate_named_range_bounds(target: XlsxNamedRange) -> None:
    if (
        target.max_row <= SPREADSHEET_MAX_ROWS
        and target.max_column <= SPREADSHEET_MAX_COLUMNS
    ):
        return
    message = f"Template target {target.reference!r} exceeds XLSX sheet bounds"
    raise IncompatibleTemplateRefError(
        message,
        context={
            "max_columns": SPREADSHEET_MAX_COLUMNS,
            "max_rows": SPREADSHEET_MAX_ROWS,
            "reference": target.reference,
        },
    )


def _validate_template_plan(plan: DocumentPlan) -> None:
    notification = Notification()
    for worksheet in plan.worksheets:
        for overlap in worksheet.overlaps:
            notification.add(
                f"Blocks {overlap.first} and {overlap.second} overlap",
                path=f'worksheet["{worksheet.name}"].{overlap.first}',
                code="block_overlap",
                context={
                    "first": overlap.first,
                    "second": overlap.second,
                    "worksheet": worksheet.name,
                },
            )
    notification.raise_if_errors("Spreadsheet structural validation failed")


def _validate_compiled_targets(  # noqa: C901
    compiled: SpreadsheetIR,
    targets: Sequence[XlsxTableTarget],
) -> None:
    notification = Notification()
    target_paths = {(item.worksheet_index, item.path) for item in targets}
    for target in targets:
        worksheet = compiled.worksheets[target.worksheet_index]
        table = _compiled_table(worksheet, target.path)
        row_count = table.rows.row_count
        target_too_short = (
            row_count is not None
            and not target.repeat
            and row_count > target.range.rows
        )
        if target_too_short:
            message = f"Template target {target.reference!r} has too few rows"
            raise IncompatibleTemplateRefError(
                message,
                context={
                    "available_rows": target.range.rows,
                    "reference": target.reference,
                    "required_rows": row_count,
                },
            )
        occupied = _target_range(target, row_count)
        exceeded = tuple(
            dimension
            for dimension, actual, limit in (
                ("rows", occupied.end.row, SPREADSHEET_MAX_ROWS),
                ("columns", occupied.end.column, SPREADSHEET_MAX_COLUMNS),
            )
            if actual > limit
        )
        if exceeded:
            notification.add(
                "Template target exceeds spreadsheet sheet bounds",
                path=f'worksheet["{worksheet.name}"].{target.path}',
                code="sheet_bounds_exceeded",
                context={"dimensions": exceeded, "reference": target.reference},
            )
        for placement in worksheet.placements:
            if (
                target.worksheet_index,
                placement.path,
            ) in target_paths or placement.occupied is None:
                continue
            if placement.kind in {
                SpreadsheetBlockKind.SPACER,
                SpreadsheetBlockKind.STACK,
            }:
                continue
            if occupied.intersects(placement.occupied):
                notification.add(
                    f"Blocks {target.path} and {placement.path} overlap",
                    path=f'worksheet["{worksheet.name}"].{target.path}',
                    code="block_overlap",
                    context={
                        "first": target.path,
                        "second": placement.path,
                        "worksheet": worksheet.name,
                    },
                )
    notification.raise_if_errors("Spreadsheet structural validation failed")


def _materialize_target_rows(
    compiled: SpreadsheetIR,
    targets: Sequence[XlsxTableTarget],
) -> SpreadsheetIR:
    paths_by_worksheet: dict[int, set[str]] = {}
    for target in targets:
        paths_by_worksheet.setdefault(target.worksheet_index, set()).add(target.path)
    worksheets = tuple(
        _materialize_worksheet_targets(
            worksheet,
            paths_by_worksheet.get(worksheet_index, set()),
        )
        for worksheet_index, worksheet in enumerate(compiled.worksheets)
    )
    return dataclasses.replace(compiled, worksheets=worksheets)


def _materialize_worksheet_targets(
    worksheet: SpreadsheetWorksheetIR,
    target_paths: set[str],
) -> SpreadsheetWorksheetIR:
    tabular_paths = (
        placement.path
        for placement in worksheet.placements
        if placement.kind in {SpreadsheetBlockKind.TABLE, SpreadsheetBlockKind.MATRIX}
    )
    tables = tuple(
        dataclasses.replace(table, rows=table.rows.materialized())
        if path in target_paths
        else table
        for path, table in zip(tabular_paths, worksheet.tables, strict=True)
    )
    return dataclasses.replace(worksheet, tables=tables)


def _compiled_table(worksheet: SpreadsheetWorksheetIR, path: str) -> SpreadsheetTableIR:
    placements = tuple(
        placement
        for placement in worksheet.placements
        if placement.kind in {SpreadsheetBlockKind.TABLE, SpreadsheetBlockKind.MATRIX}
    )
    for placement, table in zip(placements, worksheet.tables, strict=True):
        if placement.path == path:
            return table
    message = f"Compiled template target path {path!r} was not found"
    raise TemplateError(message)


def _target_range(target: XlsxTableTarget, row_count: int | None) -> CellRange:
    max_row = target.range.max_row
    if target.repeat and row_count is not None:
        max_row = target.range.min_row + max(row_count, 1) * target.range.rows - 1
    return CellRange(
        start=CellAddress(target.range.min_row, target.range.min_column),
        end=CellAddress(max_row, target.range.max_column),
    )


def _validate_target_overlaps(targets: Sequence[XlsxTableTarget]) -> None:
    for index, first in enumerate(targets):
        for second in targets[index + 1 :]:
            if _targets_intersect(first.range, second.range):
                message = (
                    f"Template targets {first.reference!r} and "
                    f"{second.reference!r} overlap"
                )
                raise IncompatibleTemplateRefError(message)


def _targets_intersect(first: XlsxNamedRange, second: XlsxNamedRange) -> bool:
    rows_intersect = first.min_row <= second.max_row and second.min_row <= first.max_row
    columns_intersect = (
        first.min_column <= second.max_column and second.min_column <= first.max_column
    )
    return first.sheet == second.sheet and rows_intersect and columns_intersect


__all__ = (
    "XlsxDefinedName",
    "XlsxNamedRange",
    "XlsxTableTarget",
    "XlsxTemplateCompiler",
    "XlsxTemplateContext",
    "XlsxTemplateInspector",
)
