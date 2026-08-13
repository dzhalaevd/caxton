from __future__ import annotations

import dataclasses
from io import BytesIO

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.formatting.rule import Rule  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment as OpenpyxlAlignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.styles.differential import (  # type: ignore[import-untyped]
    DifferentialStyle,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.filters import AutoFilter  # type: ignore[import-untyped]
from openpyxl.worksheet.table import (  # type: ignore[import-untyped]
    Table,
    TableColumn,
    TableStyleInfo,
)
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from caxton._internal.backends._xlsx_formats import number_format
from caxton._internal.const import _SEMANTIC_FEATURES
from caxton._internal.formulas import lower_excel_formula
from caxton._internal.rendering import run_backend
from caxton.core.formatting import BorderLine, FontStyle, Style
from caxton.core.ir import (
    SPREADSHEET_IR_VERSION,
    SpreadsheetColumnIR,
    SpreadsheetIR,
    SpreadsheetTableIR,
    SpreadsheetTextIR,
)
from caxton.core.models import AggregateFunction, DocumentKind
from caxton.core.protocols import OutputSink
from caxton.core.rendering import (
    ExecutionMode,
    RenderContext,
    RendererCapabilities,
    RendererDescriptor,
    RenderResult,
    WorkbookOperation,
)
from caxton.core.types import Link


class OpenpyxlRenderer:
    """Bundled create-new XLSX renderer for the current Spreadsheet IR."""

    descriptor = RendererDescriptor(
        name="openpyxl",
        version="1.0",
        formats=frozenset(("xlsx",)),
        mime_types=frozenset(
            ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",),
        ),
        extensions=frozenset((".xlsx",)),
        capabilities=RendererCapabilities(
            ir_versions={
                DocumentKind.SPREADSHEET: frozenset((SPREADSHEET_IR_VERSION,)),
            },
            features=_SEMANTIC_FEATURES
            | frozenset(
                (
                    "alignment",
                    "aggregation",
                    "autofilter",
                    "auto_width",
                    "column_width",
                    "conditional_format",
                    "display_format",
                    "explicit_anchor",
                    "flow_layout",
                    "formula",
                    "grouping",
                    "freeze_panes",
                    "native_table",
                    "matrix",
                    "merge_cells",
                    "spacer",
                    "stack",
                    "table",
                    "text",
                    "style",
                    "totals",
                ),
            ),
            workbook_operations=frozenset(
                (WorkbookOperation.CREATE_NEW_WORKBOOK,),
            ),
        ),
    )

    def render(
        self,
        document: SpreadsheetIR,
        sink: OutputSink,
        context: RenderContext,
    ) -> RenderResult:
        """Materialize Spreadsheet IR as XLSX bytes.

        Returns:
            Metadata describing the written artifact.
        """
        return run_backend(
            lambda: self._render(document, sink, context),
            message="OpenPyXL failed to render the spreadsheet",
            backend=self.descriptor.name,
        )

    def _render(
        self,
        document: SpreadsheetIR,
        sink: OutputSink,
        context: RenderContext,
    ) -> RenderResult:
        workbook = Workbook()
        for index, worksheet_ir in enumerate(document.worksheets):
            worksheet = (
                workbook.active
                if index == 0
                else workbook.create_sheet(worksheet_ir.name)
            )
            worksheet.title = worksheet_ir.name
            if worksheet_ir.freeze is not None:
                worksheet.freeze_panes = worksheet.cell(
                    row=worksheet_ir.freeze.rows + 1,
                    column=worksheet_ir.freeze.columns + 1,
                )
            for text in worksheet_ir.texts:
                _render_text(worksheet, text)
            for table in worksheet_ir.tables:
                _render_table(worksheet, table)
        buffer = BytesIO()
        workbook.save(buffer)
        payload = buffer.getvalue()
        bytes_written = sink.write(payload)
        return RenderResult(
            format=context.format,
            mime_type=next(iter(self.descriptor.mime_types)),
            renderer=self.descriptor.name,
            bytes_written=bytes_written,
            execution_mode=ExecutionMode.STANDARD,
            execution_plan="standard",
        )


def _render_text(worksheet: Worksheet, text: SpreadsheetTextIR) -> None:
    cell = worksheet.cell(
        row=text.anchor.row,
        column=text.anchor.column,
        value=text.text,
    )
    _apply_style(cell, text.style)
    if text.span > 1:
        worksheet.merge_cells(
            start_row=text.anchor.row,
            start_column=text.anchor.column,
            end_row=text.anchor.row,
            end_column=text.anchor.column + text.span - 1,
        )


def _render_table(  # noqa: C901
    worksheet: Worksheet,
    table: SpreadsheetTableIR,
) -> None:
    header_row = table.anchor.row
    start_column = table.anchor.column
    for column in table.columns:
        cell = worksheet.cell(
            row=header_row,
            column=start_column + column.offset,
            value=column.title,
        )
        _apply_style(cell, table.header_style)
        if column.width_hint is not None:
            letter = get_column_letter(start_column + column.offset)
            worksheet.column_dimensions[letter].width = column.width_hint

    last_row = header_row
    widths = [len(column.title) for column in table.columns]
    for row in table.rows:
        physical_row = header_row + row.index + 1
        last_row = physical_row
        for column, value in zip(table.columns, row.values, strict=True):
            widths[column.offset] = max(widths[column.offset], _display_width(value))
            cell_value = (
                value
                if column.formula is None
                else lower_excel_formula(
                    column.formula,
                    current_row=physical_row,
                )
            )
            cell = worksheet.cell(
                row=physical_row,
                column=start_column + column.offset,
                value=cell_value,
            )
            _style_cell(cell, column)

    for cell_range in table.merges:
        worksheet.merge_cells(
            start_row=cell_range.start.row,
            start_column=cell_range.start.column,
            end_row=cell_range.end.row,
            end_column=cell_range.end.column,
        )

    for column in table.columns:
        if column.auto_width:
            letter = get_column_letter(start_column + column.offset)
            worksheet.column_dimensions[letter].width = min(
                80,
                max(1, widths[column.offset] + 2),
            )

    _write_footer(worksheet, table, header_row, last_row, start_column)
    _add_conditional_formats(worksheet, table, header_row, last_row, start_column)
    if table.autofilter and table.name is None and last_row > header_row:
        end_column = start_column + len(table.columns) - 1
        worksheet.auto_filter.ref = (
            f"{get_column_letter(start_column)}{header_row}:"
            f"{get_column_letter(end_column)}{last_row}"
        )

    if table.name is not None:
        end_column = start_column + len(table.columns) - 1
        reference = (
            f"{get_column_letter(start_column)}{header_row}:"
            f"{get_column_letter(end_column)}{last_row}"
        )
        native_table = Table(
            displayName=table.name,
            ref=reference,
            autoFilter=AutoFilter(ref=reference) if table.autofilter else None,
            tableColumns=tuple(
                TableColumn(id=index, name=column.title)
                for index, column in enumerate(table.columns, start=1)
            ),
        )
        native_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(native_table)


def _style_cell(cell: Cell, column: SpreadsheetColumnIR) -> None:
    _apply_style(cell, column.style)
    cell.number_format = number_format(column)
    if isinstance(column.semantic_type, Link) and cell.value is not None:
        cell.hyperlink = str(cell.value)


def _apply_style(cell: Cell, style: Style) -> None:
    font, fill, border, alignment = _openpyxl_style(style)
    cell.font = font
    cell.fill = fill
    cell.border = border
    cell.alignment = alignment


def _openpyxl_style(
    style: Style,
) -> tuple[Font, PatternFill, Border, OpenpyxlAlignment]:
    font_style = style.font
    font = Font(
        name=None if font_style is None else font_style.name,
        size=None if font_style is None else font_style.size,
        bold=None if font_style is None else font_style.bold,
        italic=None if font_style is None else font_style.italic,
        underline=_underline(font_style),
        color=None if font_style is None else _argb(font_style.color),
    )
    fill = PatternFill()
    if style.fill is not None:
        fill = PatternFill(fill_type="solid", fgColor=_argb(style.fill.color))
    border = Border()
    if style.border is not None:
        border = Border(
            top=_side(style.border.top),
            right=_side(style.border.right),
            bottom=_side(style.border.bottom),
            left=_side(style.border.left),
        )
    alignment_style = style.alignment
    alignment = OpenpyxlAlignment(
        horizontal=(
            None
            if alignment_style is None or alignment_style.horizontal is None
            else alignment_style.horizontal.value
        ),
        vertical=(
            None
            if alignment_style is None or alignment_style.vertical is None
            else alignment_style.vertical.value
        ),
        wrap_text=None if alignment_style is None else alignment_style.wrap_text,
    )
    return font, fill, border, alignment


def _underline(font_style: FontStyle | None) -> str | None:
    if font_style is None or font_style.underline is None:
        return None
    return "single" if font_style.underline else None


def _side(line: BorderLine | None) -> Side:
    if line is None:
        return Side()
    return Side(style=line.style.value, color=_argb(line.color))


def _argb(color: str | None) -> str | None:
    return None if color is None else f"FF{color.removeprefix('#')}"


def _write_footer(
    worksheet: Worksheet,
    table: SpreadsheetTableIR,
    header_row: int,
    last_row: int,
    start_column: int,
) -> None:
    footer = table.footer
    if footer is None:
        return
    footer_row = last_row + 1
    label = worksheet.cell(
        row=footer_row,
        column=start_column + footer.label_column_offset,
        value=footer.label,
    )
    _apply_style(label, footer.style)
    for item in footer.items:
        column = start_column + item.column_offset
        value: object = 0
        if last_row > header_row:
            letter = get_column_letter(column)
            function = _AGGREGATES[item.function]
            value = f"={function}({letter}{header_row + 1}:{letter}{last_row})"
        cell = worksheet.cell(row=footer_row, column=column, value=value)
        _apply_style(cell, footer.style)
        effective = dataclasses.replace(
            table.columns[item.column_offset],
            display_format=(
                footer.style.display_format
                or table.columns[item.column_offset].display_format
            ),
        )
        cell.number_format = number_format(effective)


def _add_conditional_formats(
    worksheet: Worksheet,
    table: SpreadsheetTableIR,
    header_row: int,
    last_row: int,
    start_column: int,
) -> None:
    if last_row == header_row:
        return
    start = f"{get_column_letter(start_column)}{header_row + 1}"
    end = f"{get_column_letter(start_column + len(table.columns) - 1)}{last_row}"
    for item in table.rules:
        font, fill, border, _ = _openpyxl_style(item.style)
        rule = Rule(
            type="expression",
            dxf=DifferentialStyle(font=font, fill=fill, border=border),
            formula=[
                lower_excel_formula(
                    item.condition,
                    current_row=header_row + 1,
                ).removeprefix("="),
            ],
        )
        worksheet.conditional_formatting.add(f"{start}:{end}", rule)


_AGGREGATES = {
    AggregateFunction.SUM: "SUM",
    AggregateFunction.AVG: "AVERAGE",
    AggregateFunction.MIN: "MIN",
    AggregateFunction.MAX: "MAX",
    AggregateFunction.COUNT: "COUNT",
}


def _display_width(value: object) -> int:
    return 0 if value is None else len(str(value))


__all__ = ("OpenpyxlRenderer",)
