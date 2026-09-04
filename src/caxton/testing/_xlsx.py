from __future__ import annotations

from io import BytesIO
from typing import cast

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from ._artifact import (
    ArtifactCell,
    ArtifactColumn,
    ArtifactConditionalFormat,
    ArtifactTable,
    ArtifactWorksheet,
    SpreadsheetArtifact,
)
from ._errors import ArtifactInspectionError


def inspect_xlsx(payload: bytes, *, source: str) -> SpreadsheetArtifact:
    """Read XLSX bytes into backend-neutral artifact value objects.

    Returns:
        An immutable observation of the workbook.

    Raises:
        ArtifactInspectionError: If the XLSX package cannot be read.
        ModuleNotFoundError: If a missing transitive module prevents import.
    """
    try:
        workbook = _load_workbook(payload)
    except ModuleNotFoundError as error:
        if error.name != "openpyxl":
            raise
        message = "OpenPyXL runtime dependency is required for XLSX inspection"
        raise ArtifactInspectionError(
            message,
            context={"format": "xlsx", "source": source},
        ) from error
    except Exception as error:
        message = f"Failed to inspect XLSX artifact from {source}"
        raise ArtifactInspectionError(
            message,
            context={"format": "xlsx", "source": source},
        ) from error
    try:
        worksheets = tuple(_inspect_worksheet(item) for item in workbook.worksheets)
        return SpreadsheetArtifact(format="xlsx", worksheets=worksheets)
    finally:
        workbook.close()


def _load_workbook(payload: bytes) -> Workbook:
    from openpyxl import load_workbook  # noqa: PLC0415

    return load_workbook(BytesIO(payload))


def _inspect_worksheet(worksheet: Worksheet) -> ArtifactWorksheet:
    cells = tuple(
        _inspect_cell(cell)
        for row in worksheet.iter_rows()
        for cell in row
        if _is_observed(cell)
    )
    columns = tuple(
        ArtifactColumn(letter=letter, width=dimension.width)
        for letter, dimension in sorted(
            worksheet.column_dimensions.items(),
            key=_column_dimension_key,
        )
    )
    tables = tuple(
        _inspect_table(worksheet.tables[name]) for name in sorted(worksheet.tables)
    )
    merged_ranges = tuple(
        sorted(str(cell_range) for cell_range in worksheet.merged_cells.ranges)
    )
    return ArtifactWorksheet(
        name=worksheet.title,
        cells=cells,
        columns=columns,
        tables=tables,
        merged_ranges=merged_ranges,
        freeze_panes=(
            None if worksheet.freeze_panes is None else str(worksheet.freeze_panes)
        ),
        autofilter=worksheet.auto_filter.ref,
        conditional_formats=_inspect_conditional_formats(worksheet),
    )


def _is_observed(cell: Cell | MergedCell) -> bool:
    return cell.value is not None or cell.has_style or cell.hyperlink is not None


def _inspect_cell(cell: Cell | MergedCell) -> ArtifactCell:
    formula = str(cell.value) if cell.data_type == "f" else None
    hyperlink = cell.hyperlink
    hyperlink_target = None
    if hyperlink is not None:
        hyperlink_target = hyperlink.target or hyperlink.location
    return ArtifactCell(
        address=cell.coordinate,
        value=cell.value,
        formula=formula,
        number_format=cell.number_format,
        alignment=cell.alignment.horizontal,
        hyperlink=hyperlink_target,
        bold=bool(cell.font.bold),
        font_name=cell.font.name,
        font_size=None if cell.font.sz is None else float(cell.font.sz),
        font_color=_color(cell.font.color),
        fill_color=_color(cell.fill.fgColor) if cell.fill.fill_type else None,
        border_bottom=(
            None if cell.border.bottom is None else cell.border.bottom.style
        ),
    )


def _inspect_conditional_formats(
    worksheet: Worksheet,
) -> tuple[ArtifactConditionalFormat, ...]:
    inspected: list[ArtifactConditionalFormat] = []
    for conditional in worksheet.conditional_formatting:
        for rule in conditional.rules:
            differential = rule.dxf
            inspected.append(
                ArtifactConditionalFormat(
                    cell_range=str(conditional.sqref),
                    formulae=tuple(rule.formula or ()),
                    font_color=(
                        None
                        if differential is None or differential.font is None
                        else _color(differential.font.color)
                    ),
                    fill_color=(
                        None
                        if differential is None or differential.fill is None
                        else _fill_color(differential.fill)
                    ),
                ),
            )
    return tuple(inspected)


def _color(value: object) -> str | None:  # noqa: C901, WPS212
    if value is None:
        return None
    color_type = getattr(value, "type", None)
    if color_type == "theme":
        theme = getattr(value, "theme", None)
        return None if theme is None else f"theme:{theme}"
    if color_type == "indexed":
        indexed = getattr(value, "indexed", None)
        return None if indexed is None else f"indexed:{indexed}"
    if color_type == "auto":
        return "auto"
    if color_type != "rgb":
        return None
    rgb = getattr(value, "rgb", None)
    if not isinstance(rgb, str):
        return None
    normalized = rgb[-6:].upper()
    return f"#{normalized}"


def _fill_color(fill: object) -> str | None:
    foreground = _color(getattr(fill, "fgColor", None))
    if getattr(fill, "patternType", None) == "solid":
        return foreground or _color(getattr(fill, "bgColor", None))
    if foreground not in {None, "#000000"}:
        return foreground
    return _color(getattr(fill, "bgColor", None))


def _column_index(letter: str) -> int:
    from openpyxl.utils import column_index_from_string  # noqa: PLC0415

    return int(column_index_from_string(letter))


def _column_dimension_key(item: tuple[str, object]) -> int:
    return _column_index(item[0])


def _inspect_table(table: Table) -> ArtifactTable:
    from openpyxl.utils import range_boundaries  # noqa: PLC0415

    boundaries = range_boundaries(table.ref)
    if not all(isinstance(boundary, int) for boundary in boundaries):
        message = f"Table {table.displayName!r} has a non-cell range"
        raise ArtifactInspectionError(message, context={"range": table.ref})
    min_column, min_row, max_column, max_row = cast(
        "tuple[int, int, int, int]",
        boundaries,
    )
    column_titles = tuple(column.name for column in table.tableColumns)
    expected_columns = max_column - min_column + 1
    if len(column_titles) != expected_columns:
        message = (
            f"Table {table.displayName!r} declares {len(column_titles)} columns "
            f"over a {expected_columns}-column range"
        )
        raise ArtifactInspectionError(
            message,
            context={
                "table": table.displayName,
                "range": table.ref,
                "declared_columns": len(column_titles),
                "range_columns": expected_columns,
            },
        )
    return ArtifactTable(
        name=table.displayName,
        cell_range=table.ref,
        column_titles=column_titles,
        row_count=max(0, max_row - min_row),
        autofilter=table.autoFilter is not None,
    )


__all__ = ("inspect_xlsx",)
