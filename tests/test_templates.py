from __future__ import annotations

import dataclasses
from collections.abc import Iterator
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName

from caxton import (  # noqa: WPS347
    MissingTemplateRefError,
    TemplateError,
    TemplateFormatError,
    UnsupportedFeatureError,
    ValidationError,
    integer,
    ref,
    render,
    repeat,
    sheet,
    spreadsheet,
    table,
    template,
    text,
    write,
)  # noqa: WPS347
from caxton.api import xlsx
from caxton.core.models import SpreadsheetDocument


@dataclasses.dataclass(frozen=True)
class UnsupportedExtension:
    namespace: str = "example"
    required_capabilities: frozenset[str] = frozenset(("example_feature",))


def _rendered(document: SpreadsheetDocument) -> bytes:
    payload = render(document).data
    assert payload is not None
    return payload


def _template_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Report"
    worksheet["A1"] = "Existing title"
    worksheet["A3"] = "template"
    worksheet["B3"] = '=A3&"!"'
    worksheet["A3"].fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
    workbook.defined_names.add(
        DefinedName("report_data", attr_text="'Report'!$A$3:$B$6"),
    )
    workbook.defined_names.add(
        DefinedName("report_row", attr_text="'Report'!$A$3:$B$3"),
    )
    workbook.save(path)


def test_template_detects_path_format_and_is_immutable(tmp_path: Path) -> None:
    source = tmp_path / "report.xlsx"
    _template_workbook(source)

    specification = template(source)

    assert specification.format == "xlsx"
    assert specification.source == str(source)
    with pytest.raises(dataclasses.FrozenInstanceError):
        specification.format = "docx"  # type: ignore[misc]


def test_template_bytes_require_an_explicit_or_detectable_format() -> None:
    with pytest.raises(TemplateFormatError, match="detect"):
        template(b"not an office document")

    assert template(b"not an office document", format="xlsx").format == "xlsx"


def test_explicit_format_accepts_an_extensionless_source(tmp_path: Path) -> None:
    source = tmp_path / "corporate-template"

    assert template(source, format="xlsx").format == "xlsx"


def test_template_target_requires_a_document_template() -> None:
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "Ada"}],
                columns=(text(id="name", source="name"),),
                into=ref("data"),
            ),
        ),
    )

    with pytest.raises(ValidationError, match="requires a document template"):
        render(document)


def test_named_range_write_preserves_template_and_source(tmp_path: Path) -> None:
    source = tmp_path / "template.xlsx"
    _template_workbook(source)
    original = source.read_bytes()
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "Ada", "count": 2}, {"name": "Grace", "count": 3}],
                columns=(
                    text(id="name", source="name"),
                    integer(id="count", source="count"),
                ),
                into=ref("report_data"),
            ),
        ),
        template=template(source),
    )

    result = render(document)

    assert result.renderer == "openpyxl-template"
    assert source.read_bytes() == original
    assert result.data is not None
    workbook = load_workbook(BytesIO(result.data))
    worksheet = workbook["Report"]
    assert worksheet["A1"].value == "Existing title"
    assert worksheet["A3"].value == "Ada"
    assert worksheet["B3"].value == 2
    assert worksheet["A4"].value == "Grace"
    assert worksheet["B4"].value == 3
    assert worksheet["A3"].fill.fgColor.rgb == "FF00FF00"


def test_bundled_template_preserves_untouched_formula_cells() -> None:
    source = (
        Path(__file__).parents[1]
        / "example"
        / "template"
        / "assets"
        / "monthly_sales_template.xlsx"
    )
    original = source.read_bytes()
    document = spreadsheet(
        sheet(
            "Monthly Report",
            table(
                source=[{"product": "Coffee", "quantity": 4}],
                columns=(
                    text(id="product", source="product"),
                    integer(id="quantity", source="quantity"),
                ),
                into=ref("report_data"),
            ),
        ),
        template=template(source),
    )

    workbook = load_workbook(BytesIO(_rendered(document)), data_only=False)

    assert workbook["Monthly Report"]["A8"].value == "Coffee"
    assert workbook["Monthly Report"]["B8"].value == 4
    assert workbook["Monthly Report"]["F8"].value == ('=IF(COUNTA(A8:E8)=0,"",D8*E8)')
    assert source.read_bytes() == original


def test_worksheet_scoped_name_wins_for_its_sheet(tmp_path: Path) -> None:
    source = tmp_path / "scoped.xlsx"
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "First"
    second = workbook.create_sheet("Second")
    first.defined_names.add(
        DefinedName("target", attr_text="'First'!$A$2", localSheetId=0),
    )
    second.defined_names.add(
        DefinedName("target", attr_text="'Second'!$C$4", localSheetId=1),
    )
    workbook.save(source)
    document = spreadsheet(
        sheet(
            "First",
            table(
                source=[{"value": "one"}],
                columns=(text(id="value", source="value"),),
                into=ref("target"),
            ),
        ),
        sheet(
            "Second",
            table(
                source=[{"value": "two"}],
                columns=(text(id="value", source="value"),),
                into=ref("target"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(_rendered(document)))

    assert rendered["First"]["A2"].value == "one"
    assert rendered["Second"]["C4"].value == "two"


def test_missing_reference_leaves_target_untouched(tmp_path: Path) -> None:
    source = tmp_path / "template.xlsx"
    target = tmp_path / "output.xlsx"
    _template_workbook(source)
    target.write_bytes(b"existing")
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "Ada"}],
                columns=(text(id="name", source="name"),),
                into=ref("missing"),
            ),
        ),
        template=template(source),
    )

    with pytest.raises(MissingTemplateRefError, match="missing"):
        write(document, target)

    assert target.read_bytes() == b"existing"


def test_capability_preflight_leaves_target_untouched(tmp_path: Path) -> None:
    source = tmp_path / "template.xlsx"
    target = tmp_path / "output.xlsx"
    _template_workbook(source)
    target.write_bytes(b"existing")

    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "Ada"}],
                columns=(text(id="name", source="name"),),
                into=ref("report_data"),
            ),
        ),
        template=template(source, extensions=(UnsupportedExtension(),)),
    )

    with pytest.raises(UnsupportedFeatureError, match="lacks required"):
        write(document, target)

    assert target.read_bytes() == b"existing"


def test_repeat_copies_styles_formulas_and_merged_cells(tmp_path: Path) -> None:
    source = tmp_path / "repeat.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Report"
    worksheet["A2"] = "template"
    worksheet["B2"] = '=A2&"!"'
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    worksheet.merge_cells("C2:D2")
    worksheet["C2"] = "merged"
    workbook.defined_names.add(
        DefinedName("report_row", attr_text="'Report'!$A$2:$D$2"),
    )
    workbook.save(source)
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "Ada"}, {"name": "Grace"}, {"name": "Linus"}],
                columns=(text(id="name", source="name"),),
                into=repeat(ref("report_row")),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(_rendered(document)), data_only=False)
    worksheet = rendered["Report"]

    assert [worksheet.cell(row, 1).value for row in range(2, 5)] == [
        "Ada",
        "Grace",
        "Linus",
    ]
    assert [worksheet.cell(row, 2).value for row in range(2, 5)] == [
        '=A2&"!"',
        '=A3&"!"',
        '=A4&"!"',
    ]
    assert worksheet["A4"].fill.fgColor.rgb == "FFFFFF00"
    assert {str(item) for item in worksheet.merged_cells.ranges} >= {
        "C2:D2",
        "C3:D3",
        "C4:D4",
    }


def test_repeat_shifts_a_downstream_resolved_target(tmp_path: Path) -> None:
    source = tmp_path / "shift.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Report"
    workbook.defined_names.add(
        DefinedName("rows", attr_text="'Report'!$A$2:$B$2"),
    )
    workbook.defined_names.add(
        DefinedName("summary", attr_text="'Report'!$A$5"),
    )
    workbook.save(source)
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "A"}, {"name": "B"}, {"name": "C"}],
                columns=(text(id="name", source="name"),),
                into=repeat(ref("rows")),
            ),
            table(
                source=[{"value": 42}],
                columns=(integer(id="value", source="value"),),
                into=ref("summary"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(_rendered(document)))

    assert rendered["Report"]["A7"].value == 42


def test_openpyxl_hook_is_namespaced_and_scoped(tmp_path: Path) -> None:
    source = tmp_path / "hook.xlsx"
    _template_workbook(source)

    def configure(context: xlsx.OpenpyxlHookContext) -> None:
        context.native_sheet.sheet_properties.pageSetUpPr.fitToPage = True

    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"name": "Ada"}],
                columns=(text(id="name", source="name"),),
                into=ref("report_data"),
            ),
        ),
        template=template(source, extensions=(xlsx.openpyxl_hook(configure),)),
    )

    workbook = load_workbook(BytesIO(_rendered(document)))

    page_setup = workbook["Report"].sheet_properties.pageSetUpPr
    assert page_setup is not None
    assert page_setup.fitToPage


def test_missing_pivot_is_resolved_before_rows_are_consumed(tmp_path: Path) -> None:
    source = tmp_path / "pivot.xlsx"
    _template_workbook(source)
    visited = False

    def rows() -> Iterator[dict[str, str]]:
        nonlocal visited
        visited = True
        yield {"name": "Ada"}

    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=rows(),
                columns=(text(id="name", source="name"),),
                into=ref("report_data"),
            ),
        ),
        template=template(
            source,
            extensions=(
                xlsx.pivot(
                    "MissingPivot",
                    source=ref("report_data"),
                    refresh_on_open=True,
                ),
            ),
        ),
    )

    with pytest.raises(TemplateError, match="MissingPivot"):
        render(document)

    assert not visited
