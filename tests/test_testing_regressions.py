from __future__ import annotations

import dataclasses
import datetime as dt
import decimal as decimal_module
from collections.abc import Callable
from functools import partial
from io import BytesIO
from typing import Any, ClassVar, cast

import pytest
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.worksheet.table import Table, TableColumn
from openpyxl.worksheet.worksheet import Worksheet

from caxton import (
    col,
    decimal,
    image,
    sheet,
    spreadsheet,
    stack,
    table,
    template,
    text,
    title,
    when,
)
from caxton.core.errors import UnsupportedFeatureError
from caxton.core.formatting import Style
from caxton.core.models import Column, FieldRef
from caxton.core.types import SemanticType
from caxton.testing import (
    ArtifactInspectionError,
    BlockSpec,
    ConditionalRuleSpec,
    Rows,
    RowsMode,
    SpreadsheetAssertionError,
    assert_spreadsheet_equal,
    inspect_artifact,
    inspect_layout,
    inspect_spec,
)
from caxton.testing._assertions import (  # noqa: PLC2701
    _BLOCK_FIELDS,
    _COLUMN_FIELDS,
    _RULE_FIELDS,
    _SPREADSHEET_FIELDS,
    _TABLE_FIELDS,
    _WORKSHEET_FIELDS,
)
from caxton.testing._xlsx import _fill_color, _inspect_table  # noqa: PLC2701


def _callable_source(function: Callable[[object], object]) -> object:
    document = spreadsheet(
        sheet(
            "Data",
            table(source=[], columns=(text(id="value", source=function),)),
        ),
    )
    return inspect_spec(document).worksheets[0].tables[0].columns[0].source


def _constant(value: object) -> Callable[[object], object]:
    return lambda _row: value


def _return_argument(value: object, _row: object) -> object:
    return value


@dataclasses.dataclass(frozen=True)
class _Multiplier:
    factor: int

    def __call__(self, _row: object) -> int:
        return self.factor


class _ConfiguredCallable:
    def __init__(self, value: int) -> None:
        self.value = value

    def compute(self, _row: object) -> int:
        return self.value


class _NestedState:  # noqa: B903 - exercises non-dataclass object state
    def __init__(self, value: int) -> None:
        self.value = value


class _NestedCallable:
    __caxton_id__: str

    def __init__(self, value: int) -> None:
        self.state = _NestedState(value)

    def __call__(self, _row: object) -> int:
        return self.state.value


class _CyclicCallable:
    def __init__(self) -> None:
        self.state = self

    def __call__(self, _row: object) -> int:
        return 1


@pytest.mark.parametrize(
    ("first", "second"),
    [
        (partial(_return_argument, 1), partial(_return_argument, 2)),
        (
            _constant(decimal_module.Decimal("0.2")),
            _constant(decimal_module.Decimal(99)),
        ),
        (_constant(dt.date(2020, 1, 1)), _constant(dt.date(2099, 1, 1))),
        (_Multiplier(2), _Multiplier(1000)),
        (_ConfiguredCallable(1).compute, _ConfiguredCallable(2).compute),
        (_NestedCallable(1), _NestedCallable(2)),
    ],
)
def test_callable_identity_includes_supported_callable_state(
    first: Callable[[object], object],
    second: Callable[[object], object],
) -> None:
    assert _callable_source(first) != _callable_source(second)


def test_callable_identity_handles_cyclic_state() -> None:
    assert _callable_source(_CyclicCallable()) == _callable_source(_CyclicCallable())


def test_explicit_callable_identity_overrides_state() -> None:
    first = _NestedCallable(1)
    second = _NestedCallable(2)
    first.__caxton_id__ = "shared"
    second.__caxton_id__ = "shared"

    assert _callable_source(first) == _callable_source(second)


def test_block_difference_has_nested_semantic_path() -> None:
    actual = spreadsheet(sheet("Data", stack(title("Actual"))))
    expected = spreadsheet(sheet("Data", stack(title("Expected"))))

    with pytest.raises(SpreadsheetAssertionError) as captured:
        assert_spreadsheet_equal(actual, expected)

    assert [difference.path for difference in captured.value.differences] == [
        "worksheet['Data'].blocks[0].items[0].text",
    ]


def test_table_change_does_not_duplicate_block_diff() -> None:
    actual = spreadsheet(
        sheet(
            "Data",
            table(
                source=[],
                columns=(text(id="value", source="value"),),
                name="data",
                anchor="A1",
            ),
        ),
    )
    expected = spreadsheet(
        sheet(
            "Data",
            table(
                source=[],
                columns=(text(id="value", source="value"),),
                name="data",
                anchor="B1",
            ),
        ),
    )

    with pytest.raises(SpreadsheetAssertionError) as captured:
        assert_spreadsheet_equal(actual, expected)

    assert [difference.path for difference in captured.value.differences] == [
        "worksheet['Data'].table['data'].anchor",
    ]


def test_block_and_rule_order_can_be_ignored() -> None:
    positive = when(col("value") > 0, style=Style(fill="#00FF00"))
    negative = when(col("value") < 0, style=Style(fill="#FF0000"))
    actual = spreadsheet(
        sheet(
            "Data",
            title("First"),
            title("Second"),
            table(
                source=[],
                columns=(decimal(id="value", source="value"),),
                rules=(positive, negative),
            ),
        ),
    )
    expected = spreadsheet(
        sheet(
            "Data",
            title("Second"),
            title("First"),
            table(
                source=[],
                columns=(decimal(id="value", source="value"),),
                rules=(negative, positive),
            ),
        ),
    )

    assert_spreadsheet_equal(actual, expected, check_order=False)


def test_duplicate_keyed_items_are_matched_without_order() -> None:
    actual = spreadsheet(
        sheet(
            "Data",
            table(
                source=[],
                columns=(text(id="value", source="value").titled("First"),),
                name="same",
            ),
            table(
                source=[],
                columns=(text(id="value", source="value").titled("Second"),),
                name="same",
            ),
        ),
    )
    expected = spreadsheet(
        sheet(
            "Data",
            table(
                source=[],
                columns=(text(id="value", source="value").titled("Second"),),
                name="same",
            ),
            table(
                source=[],
                columns=(text(id="value", source="value").titled("First"),),
                name="same",
            ),
        ),
    )

    assert_spreadsheet_equal(actual, expected, check_order=False)


@pytest.mark.parametrize(
    ("spec_type", "compared_fields"),
    [
        (type(inspect_spec(spreadsheet())), _SPREADSHEET_FIELDS),
        (
            type(inspect_spec(spreadsheet(sheet("Data"))).worksheets[0]),
            _WORKSHEET_FIELDS,
        ),
    ],
)
def test_top_level_comparators_cover_every_spec_field(
    spec_type: type[object],
    compared_fields: frozenset[str],
) -> None:
    fields = dataclasses.fields(cast("Any", spec_type))
    assert compared_fields == {field.name for field in fields}


def test_table_and_column_comparators_cover_every_spec_field() -> None:
    document = spreadsheet(
        sheet(
            "Data",
            table(source=[], columns=(text(id="value", source="value"),)),
        ),
    )
    inspected_table = inspect_spec(document).worksheets[0].tables[0]

    assert {
        field.name for field in dataclasses.fields(type(inspected_table))
    } == _TABLE_FIELDS
    assert {
        field.name for field in dataclasses.fields(type(inspected_table.columns[0]))
    } == _COLUMN_FIELDS
    assert {field.name for field in dataclasses.fields(BlockSpec)} == _BLOCK_FIELDS
    assert {
        field.name for field in dataclasses.fields(ConditionalRuleSpec)
    } == _RULE_FIELDS


def test_rows_normalizes_string_modes_without_consuming_none_scope() -> None:
    scope = Rows("none")  # type: ignore[arg-type]
    document = spreadsheet(
        sheet(
            "Data",
            table(
                source=[{"value": 1}],
                columns=(decimal(id="value", source="value"),),
                name="data",
            ),
        ),
    )

    inspected = inspect_layout(document, rows=scope)

    assert scope.mode is RowsMode.NONE
    assert not inspected.worksheet("Data").table("data").rows


def test_rows_rejects_unknown_string_mode() -> None:
    with pytest.raises(ValueError, match="not a valid RowsMode"):
        Rows("everything")  # type: ignore[arg-type]


def test_layout_rejects_template_documents_before_reading_template() -> None:
    document = spreadsheet(
        sheet("Data"),
        template=template(b"not read by layout inspection", format="xlsx"),
    )

    with pytest.raises(
        UnsupportedFeatureError,
        match="does not model template placement",
    ) as captured:
        inspect_layout(document)

    assert captured.value.context == {"template_format": "xlsx"}


def test_layout_can_preflight_a_requested_backend() -> None:
    document = spreadsheet(sheet("Data", image(b"not rendered")))

    inspect_layout(document)
    with pytest.raises(UnsupportedFeatureError, match="lacks required capabilities"):
        inspect_layout(document, backend="openpyxl")


def test_missing_layout_row_column_lists_available_columns() -> None:
    document = spreadsheet(
        sheet(
            "Data",
            table(
                source=[{"value": 1}],
                columns=(decimal(id="value", source="value"),),
                name="data",
            ),
        ),
    )
    row = (
        inspect_layout(document, rows=Rows.all()).worksheet("Data").table("data").row(0)
    )

    with pytest.raises(KeyError, match="available columns: 'value'"):
        row["missing"]


@dataclasses.dataclass(frozen=True, slots=True, init=False)
class _DerivedSemanticType(SemanticType):
    name: ClassVar[str] = "derived"
    marker: str = dataclasses.field(init=False)

    def __init__(self, marker: str) -> None:
        object.__setattr__(self, "marker", marker)


def test_semantic_type_inspection_includes_non_init_state() -> None:
    document = spreadsheet(
        sheet(
            "Data",
            table(
                source=[],
                columns=(
                    Column(
                        id="value",
                        semantic_type=_DerivedSemanticType("visible"),
                        source=FieldRef("value"),
                    ),
                ),
            ),
        ),
    )

    semantic_type = (
        inspect_spec(document).worksheets[0].tables[0].columns[0].semantic_type
    )

    assert semantic_type.parameters == {"marker": "visible"}


def _workbook_bytes(workbook: Workbook) -> bytes:
    target = BytesIO()
    workbook.save(target)
    workbook.close()
    return target.getvalue()


@pytest.mark.parametrize(
    ("color", "expected"),
    [(Color(theme=1), "theme:1"), (Color(indexed=64), "indexed:64")],
)
def test_artifact_reports_tagged_non_rgb_color(
    color: Color,
    expected: str,
) -> None:
    workbook = Workbook()
    worksheet = cast("Worksheet", workbook.active)
    worksheet["A1"] = "colored"
    worksheet["A1"].font = Font(color=color)

    cell = inspect_artifact(_workbook_bytes(workbook)).worksheets[0].cell("A1")

    assert cell.font_color == expected


def test_conditional_format_uses_solid_foreground_color() -> None:
    fill = PatternFill(
        patternType="solid",
        fgColor="FF000000",
        bgColor="FFFFFFFF",
    )

    assert _fill_color(fill) == "#000000"


def test_artifact_columns_follow_physical_column_order() -> None:
    workbook = Workbook()
    worksheet = cast("Worksheet", workbook.active)
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["AA"].width = 30

    inspected = inspect_artifact(_workbook_bytes(workbook)).worksheets[0]

    assert [column.letter for column in inspected.columns] == ["A", "B", "AA"]


def test_artifact_exposes_observed_addresses_and_used_range() -> None:
    workbook = Workbook()
    worksheet = cast("Worksheet", workbook.active)
    worksheet["B2"] = "start"
    worksheet["D4"] = "end"

    inspected = inspect_artifact(_workbook_bytes(workbook)).worksheets[0]

    assert inspected.addresses == ("B2", "D4")
    assert inspected.used_range == "B2:D4"


def test_malformed_table_column_count_fails_loudly() -> None:
    table_value = Table(
        displayName="Broken",
        ref="A1:B2",
        tableColumns=(
            TableColumn(id=1, name="first"),
            TableColumn(id=2, name="second"),
            TableColumn(id=3, name="extra"),
        ),
    )

    with pytest.raises(ArtifactInspectionError, match="declares 3 columns") as captured:
        _inspect_table(table_value)

    assert captured.value.context == {
        "declared_columns": 3,
        "range": "A1:B2",
        "range_columns": 2,
        "table": "Broken",
    }


def test_inspector_defects_are_not_relabelled_as_invalid_artifacts(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = Workbook()
    worksheet = cast("Worksheet", workbook.active)
    worksheet["A1"] = "value"
    payload = _workbook_bytes(workbook)

    def fail(_cell: object) -> object:
        raise ZeroDivisionError

    monkeypatch.setattr("caxton.testing._xlsx._inspect_cell", fail)

    with pytest.raises(ZeroDivisionError):
        inspect_artifact(payload)
