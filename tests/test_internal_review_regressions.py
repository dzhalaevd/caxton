from __future__ import annotations

import datetime as dt
import warnings
from collections.abc import Iterator
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import ClassVar

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles.numbers import BUILTIN_FORMATS_MAX_SIZE
from openpyxl.workbook.defined_name import DefinedName

from caxton import (  # noqa: WPS347
    AggregateEvaluationError,
    CaxtonTypeError,
    IncompatibleTemplateRefError,
    OutputError,
    TemplateError,
    Total,
    Totals,
    UnsupportedFeatureError,
    ValidationError,
    chart,
    col,
    datetime,
    decimal,
    field,
    integer,
    matrix,
    render,
    repeat,
    sheet,
    sheet_ref,
    slot,
    spreadsheet,
    stack,
    table,
    template,
    text,
    title,
    when,
    write,
)
from caxton._internal import operations as operations_module  # noqa: PLC2701
from caxton._internal.aggregation import (  # noqa: PLC2701
    execution as aggregation_execution,
)
from caxton._internal.aggregation.execution import read_rows  # noqa: PLC2701
from caxton._internal.data.sources import coerce_data_source  # noqa: PLC2701
from caxton._internal.resolver import BuiltinRendererResolver  # noqa: PLC2701
from caxton._internal.semantic import SemanticRowEvaluator  # noqa: PLC2701
from caxton._internal.sinks import BufferSink  # noqa: PLC2701
from caxton.api import xlsx
from caxton.core.formatting import Style, decimal_format
from caxton.core.ir import SPREADSHEET_IR_VERSION
from caxton.core.models import DocumentKind, SpreadsheetDocument
from caxton.core.protocols import OutputSink
from caxton.core.rendering import (
    RENDERER_CONTRACT_VERSION,
    RenderContext,
    RendererCapabilities,
    RendererDescriptor,
    RenderResult,
    RequiredCapabilities,
)
from caxton.errors import PerformanceWarning
from caxton.testing import Rows, inspect_artifact, inspect_layout


def _template(path: Path, reference: str, cell_range: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Report"
    workbook.defined_names.add(
        DefinedName(reference, attr_text=f"'Report'!{cell_range}"),
    )
    workbook.save(path)


@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_text_values_and_labels_remain_literal(backend: str) -> None:
    document = spreadsheet(
        sheet(
            "Report",
            title("=heading"),
            table(
                source=[{"value": "=1+1"}, {"value": "https://example.com"}],
                columns=(text(id="value", source="value", title="=header"),),
            ),
        ),
    )

    worksheet = inspect_artifact(render(document, backend=backend)).worksheet("Report")

    assert worksheet.cell("A1").value == "=heading"
    assert worksheet.cell("A1").formula is None
    assert worksheet.cell("A2").value == "=header"
    assert worksheet.cell("A2").formula is None
    assert worksheet.cell("A3").value == "=1+1"
    assert worksheet.cell("A3").formula is None
    assert worksheet.cell("A4").hyperlink is None


def test_template_text_value_remains_literal_and_clears_old_hyperlink(
    tmp_path: Path,
) -> None:
    source = tmp_path / "template.xlsx"
    _template(source, "target", "$A$2:$B$2")
    workbook = load_workbook(source)
    worksheet = workbook["Report"]
    worksheet["A2"] = "old"
    worksheet["A2"].hyperlink = "https://old.example"
    workbook.save(source)
    workbook.close()
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": "=1+1"}],
                columns=(text(id="value", source="value"),),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""), data_only=False)

    assert rendered["Report"]["A2"].value == "=1+1"
    assert rendered["Report"]["A2"].data_type == "s"
    assert rendered["Report"]["A2"].hyperlink is None


class _ChunkRenderer:
    payload: ClassVar[bytes] = b"firstsecond"

    descriptor = RendererDescriptor(
        name="chunks",
        version="1.0",
        formats=frozenset(("xlsx",)),
        mime_types=frozenset(("application/test",)),
        extensions=frozenset((".xlsx",)),
        capabilities=RendererCapabilities(
            ir_versions={
                DocumentKind.SPREADSHEET: frozenset((SPREADSHEET_IR_VERSION,))
            },
            features=frozenset(("semantic:text", "table")),
        ),
        contract_version=RENDERER_CONTRACT_VERSION,
    )

    def __init__(self, *, fail: bool = False) -> None:
        self.fail = fail

    def render(
        self,
        _document: object,
        sink: OutputSink,
        context: RenderContext,
    ) -> RenderResult:
        sink.write(b"first")
        if self.fail:
            message = "renderer failed"
            raise RuntimeError(message)
        sink.write(b"second")
        return RenderResult(
            format=context.format,
            mime_type="application/test",
            renderer="chunks",
            bytes_written=len(self.payload),
        )


def _plain_document() -> SpreadsheetDocument:
    return spreadsheet(
        sheet(
            "Report",
            table(source=[{"value": "x"}], columns=(text(source="value"),)),
        ),
    )


def test_file_output_commits_all_chunks_only_after_renderer_success(
    tmp_path: Path,
) -> None:
    target = tmp_path / "output.xlsx"
    target.write_bytes(b"original")

    with pytest.raises(RuntimeError, match="renderer failed"):
        write(
            _plain_document(),
            target,
            renderer=_ChunkRenderer(fail=True),
        )
    assert target.read_bytes() == b"original"

    write(
        _plain_document(),
        target,
        renderer=_ChunkRenderer(),
    )
    assert target.read_bytes() == _ChunkRenderer.payload


def test_seekable_buffer_is_overwritten_and_truncated() -> None:
    target = BytesIO(b"old payload that is longer")
    target.seek(7)

    write(
        _plain_document(),
        target,
        renderer=_ChunkRenderer(),
    )

    assert target.getvalue() == _ChunkRenderer.payload


def test_matrix_before_template_target_does_not_steal_binding(tmp_path: Path) -> None:
    source = tmp_path / "template.xlsx"
    _template(source, "target", "$D$4")
    document = spreadsheet(
        sheet(
            "Report",
            matrix(
                source=[{"row": "A", "column": "X", "value": 1}],
                row=field("row"),
                column=field("column"),
                value=field("value"),
                anchor="A1",
            ),
            table(
                source=[{"value": "bound"}],
                columns=(text(source="value"),),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""))

    assert rendered["Report"]["D4"].value == "bound"
    assert rendered["Report"]["A1"].value == "row"


def test_nested_template_target_uses_its_block_path(tmp_path: Path) -> None:
    source = tmp_path / "nested.xlsx"
    _template(source, "target", "$D$4")
    document = spreadsheet(
        sheet(
            "Report",
            stack(
                matrix(
                    source=[{"row": "A", "column": "X", "value": 1}],
                    row=field("row"),
                    column=field("column"),
                    value=field("value"),
                    anchor="A1",
                ),
                table(
                    source=[{"value": "nested"}],
                    columns=(text(source="value"),),
                    into=slot("target"),
                ),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""))["Report"]

    assert rendered["D4"].value == "nested"


def test_unknown_vertical_flow_remains_unknown_after_explicit_anchor() -> None:
    def rows() -> Iterator[dict[str, str]]:
        yield {"value": "x"}

    document = spreadsheet(
        sheet(
            "Report",
            table(source=rows(), columns=(text(source="value"),)),
            title("explicit", anchor="A5"),
            title("implicit"),
        ),
    )

    with pytest.raises(UnsupportedFeatureError, match="preceding block"):
        inspect_layout(document, rows=Rows.none())


def test_template_target_overlap_with_semantic_title_is_rejected(
    tmp_path: Path,
) -> None:
    source = tmp_path / "template.xlsx"
    _template(source, "target", "$A$1:$A$3")
    document = spreadsheet(
        sheet(
            "Report",
            title("Heading", anchor="A1"),
            table(
                source=[{"value": "x"}],
                columns=(text(source="value"),),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    with pytest.raises(ValidationError) as captured:
        render(document)

    assert {issue.code for issue in captured.value.issues} == {"block_overlap"}


def test_template_target_outside_xlsx_bounds_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "bounds.xlsx"
    _template(source, "target", "$A$1048577")
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": "x"}],
                columns=(text(source="value"),),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    with pytest.raises(IncompatibleTemplateRefError, match="bounds"):
        render(document)


@pytest.mark.parametrize(
    ("column", "value"),
    [
        (integer(source="value"), 1_234_567_890_123_456),
        (integer(source="value"), 12_345_678_901_234_567_890),
        (integer(source="value"), float("nan")),
        (text(source="value"), "x" * 32_768),
        (datetime(source="value"), dt.datetime(2026, 1, 1, tzinfo=dt.timezone.utc)),
    ],
    ids=(
        "integer-precision-loss",
        "integer-out-of-range",
        "non-finite-number",
        "text-too-long",
        "timezone-aware-datetime",
    ),
)
@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_unrepresentable_xlsx_values_have_stable_errors(
    backend: str,
    column: object,
    value: object,
) -> None:
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": value}],
                columns=(column,),  # type: ignore[arg-type]
            ),
        ),
    )

    with pytest.raises(UnsupportedFeatureError) as captured:
        render(document, backend=backend)

    assert captured.value.context["worksheet"] == "Report"
    assert captured.value.context["row"] == 0
    assert captured.value.context["column"] == "value"


@pytest.mark.parametrize(
    ("cell", "formula"),
    [("B5", "=A5*2"), ("B1", "=A5*2")],
)
def test_repeat_rejects_external_formula_instead_of_corrupting_it(
    tmp_path: Path,
    cell: str,
    formula: str,
) -> None:
    source = tmp_path / "repeat.xlsx"
    _template(source, "rows", "$A$2")
    workbook = load_workbook(source)
    worksheet = workbook["Report"]
    worksheet["A5"] = 2
    worksheet[cell] = formula
    workbook.save(source)
    workbook.close()
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": "A"}, {"value": "B"}],
                columns=(text(source="value"),),
                into=repeat(slot("rows")),
            ),
        ),
        template=template(source),
    )

    with pytest.raises(IncompatibleTemplateRefError, match="downstream"):
        render(document)


@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_decimal_trailing_zeroes_do_not_trigger_precision_rejection(
    backend: str,
) -> None:
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": Decimal("1.230000000000000")}],
                columns=(decimal(source="value"),),
            ),
        ),
    )

    assert render(document, backend=backend).data is not None


def test_empty_repeat_clears_original_template_block(tmp_path: Path) -> None:
    source = tmp_path / "empty-repeat.xlsx"
    _template(source, "rows", "$A$2:$B$2")
    workbook = load_workbook(source)
    worksheet = workbook["Report"]
    worksheet["A2"] = "old"
    worksheet["B2"] = "=A2"
    workbook.save(source)
    workbook.close()
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[],
                columns=(text(source="value"),),
                into=repeat(slot("rows")),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""))["Report"]

    assert rendered["A2"].value is None
    assert rendered["B2"].value is None


@pytest.mark.parametrize("row_count", [0, 1])
def test_template_target_clears_unused_old_values(
    tmp_path: Path,
    row_count: int,
) -> None:
    source = tmp_path / f"normal-{row_count}.xlsx"
    _template(source, "target", "$A$2:$A$4")
    workbook = load_workbook(source)
    worksheet = workbook["Report"]
    for row in range(2, 5):
        worksheet.cell(row, 1, f"old-{row}").hyperlink = "https://old.example"
    workbook.save(source)
    workbook.close()
    rows = [{"value": "new"}] if row_count else []
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=rows,
                columns=(text(source="value"),),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""))["Report"]

    expected_first_value = "new" if row_count else None
    assert rendered["A2"].value == expected_first_value
    assert [rendered.cell(row, 1).value for row in range(3, 5)] == [None, None]
    assert all(rendered.cell(row, 1).hyperlink is None for row in range(2, 5))


def test_template_target_clears_unused_columns(tmp_path: Path) -> None:
    source = tmp_path / "wide-target.xlsx"
    _template(source, "target", "$A$2:$C$4")
    workbook = load_workbook(source)
    worksheet = workbook["Report"]
    for row in range(2, 5):
        for column in range(1, 4):
            worksheet.cell(
                row, column, f"old-{row}-{column}"
            ).hyperlink = "https://old.example"
    worksheet["C2"] = "=1+1"
    workbook.save(source)
    workbook.close()
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": "new"}],
                columns=(text(source="value"),),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""))["Report"]

    assert rendered["A2"].value == "new"
    assert rendered["C2"].value == "=1+1"
    assert all(
        rendered.cell(row, column).value is None
        and rendered.cell(row, column).hyperlink is None
        for row in range(2, 5)
        for column in range(1, 4)
        if (row, column) not in {(2, 1), (2, 3)}
    )


@pytest.mark.parametrize("one_shot", [False, True])
def test_repeat_growth_is_checked_for_overlap(
    tmp_path: Path,
    one_shot: bool,
) -> None:
    source = tmp_path / "repeat-overlap.xlsx"
    _template(source, "rows", "$A$2")
    rows = [{"value": "r1"}, {"value": "r2"}, {"value": "r3"}]
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=(iter(rows) if one_shot else rows),
                columns=(text(source="value"),),
                into=repeat(slot("rows")),
            ),
            table(
                source=[{"value": "other"}],
                columns=(text(source="value"),),
                anchor="A4",
            ),
        ),
        template=template(source),
    )

    with pytest.raises(ValidationError, match="overlap"):
        render(document)


def test_grouped_template_height_uses_prepared_output_shape(tmp_path: Path) -> None:
    source = tmp_path / "grouped.xlsx"
    _template(source, "target", "$A$2:$B$2")
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"group": "A", "value": 1}, {"group": "A", "value": 2}],
                columns=(
                    text(source="group").grouped(),
                    integer(id="total", source=field("value").agg(sum)),
                ),
                into=slot("target"),
            ),
        ),
        template=template(source),
    )

    rendered = load_workbook(BytesIO(render(document).data or b""))["Report"]

    assert rendered["A2"].value == "A"


def test_missing_template_extension_target_precedes_grouped_source_consumption(
    tmp_path: Path,
) -> None:
    source = tmp_path / "extension.xlsx"
    _template(source, "target", "$A$2:$B$2")
    visited = False

    def rows() -> Iterator[dict[str, object]]:
        nonlocal visited
        visited = True
        yield {"group": "A", "value": 1}

    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=rows(),
                columns=(
                    text(source="group").grouped(),
                    integer(id="total", source=field("value").agg(sum)),
                ),
                into=slot("target"),
            ),
        ),
        template=template(
            source,
            extensions=(xlsx.pivot("missing", source=slot("target")),),
        ),
    )

    with pytest.raises(TemplateError):
        render(document)

    assert not visited


def test_target_table_rejects_unmaterialized_presentation_intent(
    tmp_path: Path,
) -> None:
    source = tmp_path / "target.xlsx"
    _template(source, "target", "$A$2:$A$3")
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": "x"}],
                columns=(text(source="value"),),
                into=slot("target"),
                autofilter=True,
            ),
        ),
        template=template(source),
    )

    with pytest.raises(UnsupportedFeatureError, match="target tables"):
        render(document)


@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_empty_named_table_renders_with_both_backends(backend: str) -> None:
    document = spreadsheet(
        sheet(
            "Report",
            table(source=[], columns=(text(source="value"),), name="records"),
        ),
    )

    worksheet = inspect_artifact(render(document, backend=backend)).worksheet("Report")

    assert worksheet.table("records").column_titles == ("value",)


@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_empty_named_table_keeps_footer_outside_table(backend: str) -> None:
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[],
                columns=(text(source="label"), integer(source="n")),
                name="records",
                footer=Totals(label="Total", items=(Total("n"),)),
            ),
            title("After"),
        ),
    )

    workbook = load_workbook(BytesIO(render(document, backend=backend).data or b""))
    worksheet = workbook["Report"]

    assert worksheet.tables["records"].ref == "A1:B2"
    assert worksheet["A2"].value is None
    assert worksheet["A3"].value == "Total"
    assert worksheet["B3"].value == "=SUM(B2:B2)"
    assert worksheet["A4"].value == "After"


@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_large_power_of_ten_integer_is_preserved(backend: str) -> None:
    value = 10**16
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": value}],
                columns=(integer(source="value"),),
            ),
        ),
    )

    rendered = load_workbook(BytesIO(render(document, backend=backend).data or b""))

    assert rendered["Report"]["A2"].value == value


def test_explicit_chart_sheet_reference_never_falls_back_globally() -> None:
    document = spreadsheet(
        sheet(
            "Sales",
            chart(sheet_ref("Sales").table("orders"), x="item", y="amount"),
        ),
        sheet(
            "Archive",
            table(
                source=[{"item": "x", "amount": 1}],
                columns=(text(source="item"), integer(source="amount")),
                name="orders",
            ),
        ),
    )

    with pytest.raises(ValidationError) as captured:
        render(document)

    assert {issue.code for issue in captured.value.issues} == {"table_not_found"}


def test_reused_matrix_has_reiterable_prepared_rows() -> None:
    pivot = matrix(
        source=[{"row": "A", "column": "X", "value": 1}],
        row=field("row"),
        column=field("column"),
        value=field("value"),
    )
    document = spreadsheet(sheet("First", pivot), sheet("Second", pivot))

    artifact = inspect_artifact(render(document))

    assert artifact.worksheet("First").cell("B2").value == 1
    assert artifact.worksheet("Second").cell("B2").value == 1


class _BadBool:
    def __bool__(self) -> bool:
        message = "bad truth value"
        raise RuntimeError(message)


def test_aggregate_predicate_bool_error_has_context() -> None:
    aggregate = field("value").agg(sum, where=field("include"))

    with pytest.raises(AggregateEvaluationError) as captured:
        read_rows(
            coerce_data_source([{"value": 1, "include": _BadBool()}]),
            (),
            SemanticRowEvaluator(),
            aggregates=(aggregate,),
            path="table",
        )

    assert captured.value.context["phase"] == "predicate"
    assert captured.value.context["row_index"] == 0
    assert isinstance(captured.value.__cause__, RuntimeError)


class _InvalidWriter:
    def __init__(self, result: object) -> None:
        self.result = result

    def write(self, _data: bytes) -> int:
        if isinstance(self.result, Exception):
            raise self.result
        return self.result  # type: ignore[return-value]


@pytest.mark.parametrize("result", [0, -1, True, "bad", ValueError("closed")])
def test_buffer_delivery_protocol_errors_are_output_errors(result: object) -> None:
    with pytest.raises(OutputError):
        write(
            _plain_document(),
            _InvalidWriter(result),
            renderer=_ChunkRenderer(),
        )


class _ShortOnceWriter:
    def __init__(self) -> None:
        self.calls: list[int] = []
        self.output = bytearray()

    def write(self, data: bytes) -> int:
        self.calls.append(len(data))
        accepted = 1 if len(self.calls) == 1 else len(data)
        self.output.extend(data[:accepted])
        return accepted


def test_chunk_size_recovers_after_short_write() -> None:
    writer = _ShortOnceWriter()
    payload = b"x" * (128 * 1024)

    BufferSink(writer).write(payload)

    assert bytes(writer.output) == payload
    assert writer.calls[:3] == [64 * 1024, 2, 64 * 1024]


def test_transaction_type_error_is_a_caxton_error() -> None:
    invalid_sink = object()
    with pytest.raises(CaxtonTypeError):
        operations_module._output_transaction(  # noqa: SLF001
            invalid_sink,  # type: ignore[arg-type]
        )


def test_large_buffer_warning_precedes_full_materialization(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def rows() -> Iterator[dict[str, object]]:
        yield {"group": "A", "value": 1}
        yield {"group": "A", "value": 2}
        assert any(isinstance(item.message, PerformanceWarning) for item in caught)
        yield {"group": "A", "value": 3}

    monkeypatch.setattr(
        aggregation_execution,
        "BUFFERED_ROW_WARNING_THRESHOLD",
        1,
    )
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=rows(),
                columns=(
                    text(source="group").grouped(),
                    integer(id="total", source=field("value").agg(sum)),
                ),
            ),
        ),
    )

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        render(document)


def test_resolver_filters_incompatible_candidates_before_ambiguity() -> None:
    compatible = _ChunkRenderer()
    incompatible = _ChunkRenderer()
    incompatible.descriptor = RendererDescriptor(
        name="limited",
        version="1.0",
        formats=frozenset(("xlsx",)),
        mime_types=frozenset(("application/test",)),
        extensions=frozenset((".xlsx",)),
        capabilities=RendererCapabilities(
            ir_versions={
                DocumentKind.SPREADSHEET: frozenset((SPREADSHEET_IR_VERSION,))
            },
            features=frozenset(),
        ),
    )
    required = RequiredCapabilities(
        document_kind=DocumentKind.SPREADSHEET,
        ir_versions=frozenset((SPREADSHEET_IR_VERSION,)),
        features=frozenset(("table",)),
    )

    selected = BuiltinRendererResolver((compatible, incompatible)).select(
        required,
        format_name="xlsx",
    )

    assert selected is compatible


@pytest.mark.parametrize("backend", ["xlsxwriter", "openpyxl"])
def test_conditional_style_materializes_display_format(backend: str) -> None:
    document = spreadsheet(
        sheet(
            "Report",
            table(
                source=[{"value": 1.5}],
                columns=(decimal(source="value"),),
                rules=(
                    when(
                        col("value") > 0,
                        style=Style(display_format=decimal_format(places=3)),
                    ),
                ),
            ),
        ),
    )

    payload = render(document, backend=backend).data
    workbook = load_workbook(BytesIO(payload or b""))
    rules = tuple(workbook["Report"].conditional_formatting)
    native_rules = workbook["Report"].conditional_formatting[rules[0]]

    assert native_rules[0].dxf.numFmt.formatCode == "0.000"
    if backend == "openpyxl":
        assert native_rules[0].dxf.numFmt.numFmtId >= BUILTIN_FORMATS_MAX_SIZE
